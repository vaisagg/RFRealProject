from datetime import date

import openpyxl


def update_covid_data(cases,deaths,recovered):
    f = openpyxl.load_workbook("..\Testdata\Covid_Data.xlsx")
    sheet = f["Covid"]
    mc = sheet.max_column
    mr = int(sheet.max_row)
    sheet.cell(mr + 1, 1).value = date.today()
    sheet.cell(mr + 1, 2).value = cases
    sheet.cell(mr + 1, 3).value = deaths
    sheet.cell(mr + 1, 4).value = recovered
    f.save("..\Testdata\Covid_Data.xlsx")


